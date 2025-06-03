import requests
import openpyxl
import os

API_URL = os.environ.get('API_URL', 'http://localhost:8000')
EXCEL_FILE = os.path.join(os.path.dirname(__file__), 'ejemplo_importacion.xlsx')

# Cargar el archivo Excel
wb = openpyxl.load_workbook(EXCEL_FILE)

# Importar signature
if 'signature' in wb.sheetnames:
    ws = wb['signature']
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        try:
            r = requests.post(f'{API_URL}/signature/create', json=data)
            print('Signature:', data, '->', r.status_code, r.text)
        except Exception as e:
            print('Error importando signature:', data, e)
else:
    print('No se encontró la hoja signature')

# Importar Program
if 'Program' in wb.sheetnames:
    ws = wb['Program']
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        try:
            r = requests.post(f'{API_URL}/program/create', json=data)
            print('Program:', data, '->', r.status_code, r.text)
        except Exception as e:
            print('Error importando program:', data, e)
else:
    print('No se encontró la hoja Program')
